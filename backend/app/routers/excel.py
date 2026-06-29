import io
import uuid
from typing import Annotated

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import get_current_user
from app.models.question import (
    AnswerChoice, Difficulty, Question, QuestionSource, QuestionStatus,
)
from app.models.question_bank import QuestionBank
from app.models.user import User

router = APIRouter()

COLUMNS = [
    ("题干", "content"),
    ("选项A", "option_a"),
    ("选项B", "option_b"),
    ("选项C", "option_c"),
    ("选项D", "option_d"),
    ("正确答案(A/B/C/D)", "correct_answer"),
    ("解析(可选)", "explanation"),
    ("难度(easy/medium/hard)", "difficulty"),
]

DIFFICULTY_MAP = {"基础": "easy", "进阶": "medium", "综合": "hard",
                  "easy": "easy", "medium": "medium", "hard": "hard"}


@router.get("/template", summary="下载导入模板")
async def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题目导入模板"

    headers = [col[0] for col in COLUMNS]
    ws.append(headers)

    from openpyxl.styles import Alignment, Font, PatternFill
    header_fill = PatternFill("solid", fgColor="2F4BFF")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    col_widths = [50, 30, 30, 30, 30, 12, 40, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.append([
        "以下哪个协议工作在应用层？",
        "TCP", "HTTP", "IP", "ARP",
        "B", "HTTP 是超文本传输协议，工作在应用层。", "easy",
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quizforge_import_template.xlsx"},
    )


@router.post("/{bank_id}/import", summary="导入 Excel 题目")
async def import_questions(
    bank_id: uuid.UUID,
    file: Annotated[UploadFile, File(description="Excel 文件")],
    auto_approve: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    bank = await db.get(QuestionBank, bank_id)
    if bank is None:
        raise HTTPException(404, "题库不存在")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "文件格式错误，请使用下载的 Excel 模板")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(400, "文件中没有题目数据")
    if len(rows) > 500:
        raise HTTPException(400, "单次导入不能超过 500 道题目")

    status = QuestionStatus.approved if auto_approve else QuestionStatus.pending_review
    errors = []
    saved = 0

    for row_num, row in enumerate(rows, start=2):
        if not any(row):
            continue
        try:
            content_val = str(row[0] or "").strip()
            opt_a = str(row[1] or "").strip()
            opt_b = str(row[2] or "").strip()
            opt_c = str(row[3] or "").strip()
            opt_d = str(row[4] or "").strip()
            answer = str(row[5] or "").strip().upper()
            explanation = str(row[6] or "").strip() or None
            difficulty_raw = str(row[7] or "medium").strip()

            if not content_val:
                errors.append(f"第 {row_num} 行：题干不能为空")
                continue
            if not all([opt_a, opt_b, opt_c, opt_d]):
                errors.append(f"第 {row_num} 行：四个选项不能为空")
                continue
            if answer not in ("A", "B", "C", "D"):
                errors.append(f"第 {row_num} 行：正确答案必须是 A/B/C/D，当前值：{answer!r}")
                continue

            difficulty = DIFFICULTY_MAP.get(difficulty_raw, "medium")

            q = Question(
                bank_id=bank_id,
                content=content_val,
                option_a=opt_a, option_b=opt_b, option_c=opt_c, option_d=opt_d,
                correct_answer=AnswerChoice(answer),
                explanation=explanation,
                difficulty=Difficulty(difficulty),
                source=QuestionSource.imported,
                status=status,
                created_by=current_user.id,
            )
            db.add(q)
            saved += 1

        except Exception as e:
            errors.append(f"第 {row_num} 行处理失败：{str(e)}")

    if saved > 0:
        bank.question_count += saved
        await db.commit()

    return {
        "imported": saved,
        "errors": errors[:20],
        "total_errors": len(errors),
        "detail": f"成功导入 {saved} 道题目" + (f"，{len(errors)} 行存在错误" if errors else ""),
    }


@router.get("/{bank_id}/export", summary="导出题库为 Excel")
async def export_questions(
    bank_id: uuid.UUID,
    status: QuestionStatus | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    bank = await db.get(QuestionBank, bank_id)
    if bank is None:
        raise HTTPException(404, "题库不存在")

    stmt = select(Question).where(Question.bank_id == bank_id)
    if status:
        stmt = stmt.where(Question.status == status)
    stmt = stmt.order_by(Question.created_at)
    questions = list(await db.scalars(stmt))

    if not questions:
        raise HTTPException(404, "该题库没有符合条件的题目")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = bank.title[:31]

    from openpyxl.styles import Alignment, Font, PatternFill
    headers = [col[0] for col in COLUMNS] + ["状态", "来源", "AI模型"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="2F4BFF")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    STATUS_ZH = {"approved": "已通过", "pending_review": "待审核", "rejected": "已拒绝"}
    SOURCE_ZH = {"ai_generated": "AI生成", "manual": "手动", "imported": "导入"}

    for q in questions:
        ws.append([
            q.content, q.option_a, q.option_b, q.option_c, q.option_d,
            q.correct_answer.value, q.explanation or "",
            q.difficulty.value,
            STATUS_ZH.get(q.status.value, q.status.value),
            SOURCE_ZH.get(q.source.value, q.source.value),
            q.ai_model or "",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = bank.title.replace("/", "_").replace("\\", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_title}.xlsx"},
    )
